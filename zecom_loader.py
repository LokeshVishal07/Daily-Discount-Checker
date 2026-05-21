"""
utils/zecom_loader.py

Fully dynamic ZeCom loader — works regardless of which columns move between
bi-weekly file updates. Zero hardcoded column letters or positions.

Strategy:
1. Auto-detect the real header row (highest non-null count in first 6 rows)
2. Build column labels as  "Header Name [COL_LETTER]"  so duplicates are unique
3. Return the full dataframe + two filtered lists:
      numeric_cols  → for RRP / SRP / MD Price pickers
      text_cols     → for Remarks / Exclusion pickers
4. The UI lets the user pick freely from those lists every session
"""

from __future__ import annotations
import io
import re
import logging
from typing import Optional

import pandas as pd
import openpyxl
import openpyxl.utils as ou

logger = logging.getLogger(__name__)


# ─── Public ───────────────────────────────────────────────────────────────────

def get_sheet_names(file_bytes: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    return wb.sheetnames


def load_zecom_sheet(
    file_bytes: bytes,
    sheet_name: str,
) -> tuple[pd.DataFrame | None, list[str], list[str], list[str], str]:
    """
    Load one sheet from a ZeCom file.

    Returns:
        df            — full dataframe, column names = "Header [COL_LETTER]"
        numeric_cols  — cols with numeric data  → RRP / SRP pickers
        text_cols     — cols with text data     → Remarks / Exclusion pickers
        all_cols      — every column name
        error         — '' on success
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None, [], [], [], f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return None, [], [], [], "Sheet has too few rows."

        # ── Auto-detect header row (first 6 rows, pick one with most non-null cells)
        header_idx = max(range(min(6, len(rows))),
                         key=lambda i: sum(1 for v in rows[i] if v is not None))
        raw_headers = list(rows[header_idx])
        data_rows   = [list(r) for r in rows[header_idx + 1:]]

        # ── Build unique column names  "Header Name [COL_LETTER]"
        col_names = []
        for i, h in enumerate(raw_headers):
            letter = ou.get_column_letter(i + 1)
            label  = _clean_header(h)
            col_names.append(f"{label} [{letter}]" if label else f"[{letter}]")

        df = pd.DataFrame(data_rows, columns=col_names)
        df = df.dropna(how="all").reset_index(drop=True)

        # ── Categorise columns
        numeric_cols, text_cols = [], []
        for col in df.columns:
            series    = df[col].dropna()
            converted = pd.to_numeric(series, errors="coerce")
            num_ratio = converted.notna().sum() / max(len(series), 1)
            if num_ratio > 0.5 and converted.notna().sum() >= 3:
                numeric_cols.append(col)
            elif len(series) > 0:
                text_cols.append(col)

        return df, numeric_cols, text_cols, col_names, ""

    except Exception as exc:
        logger.exception("Error loading ZeCom sheet '%s'", sheet_name)
        return None, [], [], [], str(exc)


def build_article_lookup(
    df: pd.DataFrame,
    article_col: str,
    rrp_col: str,
    srp_col: Optional[str],
    remarks_col: str,
    platform_vc_col: Optional[str],
) -> tuple[pd.DataFrame | None, str]:
    """
    Build a clean per-article lookup table:
        Article Number | RRP | SRP | remark | platform_vc
    """
    try:
        lookup = pd.DataFrame()
        lookup["Article Number"] = (
            df[article_col].astype(str).str.strip()
            .str.replace(r"\.0$", "", regex=True)
        )
        lookup["RRP"] = pd.to_numeric(df[rrp_col], errors="coerce")

        if srp_col and srp_col in df.columns:
            srp_raw = pd.to_numeric(df[srp_col], errors="coerce")
            # Where SRP is missing/zero, fall back to RRP
            lookup["SRP"] = srp_raw.where(srp_raw.notna() & (srp_raw > 0), lookup["RRP"])
        else:
            lookup["SRP"] = lookup["RRP"]

        lookup["remark"] = (
            df[remarks_col].astype(str).str.strip()
            .replace({"None": "", "nan": "", "0": ""})
            if remarks_col and remarks_col in df.columns
            else ""
        )

        lookup["platform_vc"] = (
            df[platform_vc_col].astype(str).str.strip()
            if platform_vc_col and platform_vc_col in df.columns
            else ""
        )

        # Drop blank / invalid article numbers
        lookup = lookup[
            lookup["Article Number"].notna() &
            (lookup["Article Number"] != "") &
            (lookup["Article Number"] != "nan") &
            (lookup["Article Number"].str.len() >= 3)
        ]
        lookup = lookup.drop_duplicates("Article Number").reset_index(drop=True)
        return lookup, ""

    except Exception as exc:
        logger.exception("Error building article lookup")
        return None, str(exc)


def guess_article_col(all_cols: list[str]) -> Optional[str]:
    """
    Heuristically guess the Article Number column.
    Looks for: Style#, STYLE#, Article#, PIM Article#, Color_No etc.
    """
    keywords = ["style#", "article#", "pim article", "color_no", "colour_no",
                "articleno", "style no", "sku", "item#"]
    for col in all_cols:
        col_lower = col.lower()
        if any(k in col_lower for k in keywords):
            return col
    return None


def guess_rrp_col(numeric_cols: list[str], region: str) -> Optional[str]:
    """Guess RRP column based on region name patterns."""
    region_upper = region.upper()
    # Prefer column headers containing the region + RRP
    patterns = [f"{region_upper} RRP", "RRP", "EC RRP", "retail price", "rrp"]
    for pat in patterns:
        for col in numeric_cols:
            if pat.lower() in col.lower():
                return col
    return numeric_cols[0] if numeric_cols else None


def guess_srp_col(numeric_cols: list[str], region: str) -> Optional[str]:
    """Guess SRP/MD Price column."""
    patterns = ["ec srp", "md price", "srp", "markdown", "md price"]
    for pat in patterns:
        for col in numeric_cols:
            if pat.lower() in col.lower():
                return col
    return None


def guess_remarks_col(text_cols: list[str]) -> Optional[str]:
    """Guess MP Remarks column."""
    patterns = ["mp promo rmk", "mp remarks", "exclusion", "promo rmk mp",
                "eoss promo rmk mp", "remarks"]
    for pat in patterns:
        for col in text_cols:
            if pat.lower() in col.lower():
                return col
    return None


def guess_platform_vc_col(text_cols: list[str]) -> Optional[str]:
    patterns = ["platform vc", "platform voucher"]
    for pat in patterns:
        for col in text_cols:
            if pat.lower() in col.lower():
                return col
    return None


# ─── Private ──────────────────────────────────────────────────────────────────

def _clean_header(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Remove newlines that appear in some headers
    s = re.sub(r"[\r\n]+", " ", s)
    return s.strip()
