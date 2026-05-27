"""
exporter.py — Excel report builder aligned to Discount_Check_Format.xlsx
"""
from __future__ import annotations
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
DARK_FILL  = PatternFill("solid", fgColor="1F3864")
RED_FILL   = PatternFill("solid", fgColor="C00000")
ALT_FILL   = PatternFill("solid", fgColor="F2F2F2")
CENTER     = Alignment(horizontal="center", wrap_text=True)
LEFT       = Alignment(horizontal="left")

SEV_COLORS = {
    "red":    "FFCCCC",
    "orange": "FFE5CC",
    "amber":  "FFF5CC",
    "green":  "CCFFDD",
    "grey":   "EEEEEE",
}

# Exact column order for Full Detail / Flagged Orders (matches format file)
DETAIL_COLS = [
    "region", "marketplace", "order_id", "sku", "Article Number",
    "product_name", "order_status", "rrp_used", "srp_used",
    "seller_srp_disc_pct",           # Col 10: Seller SRP Discount %
    "seller_vc_disc_pct",            # Col 11: SELLER Voucher Discount % from Remark
    "seller_end_disc_pct",           # Col 12: SELLER END DISCOUNT %
    "paid_price",                    # Col 13: Customer PAID Price
    "actual_total_disc_pct",         # Col 14: Customer Disc % from RRP
    "platform_discount_amount",      # Col 15: Platform Discount Amount
    "effective_price",               # Col 16: Effective Price (Paid + Platform Disc)
    "effective_disc_pct",            # Col 17: Effective Disc % from RRP ← used for flagging
    "remark", "rule_label", "rule_type",
    "flagged", "flag_reason", "flag_severity",
]

# Exact display names matching the format file
COL_LABELS = {
    "region":                   "Region",
    "marketplace":              "Marketplace",
    "order_id":                 "Order Id",
    "sku":                      "Sku",
    "Article Number":           "Article Number",
    "product_name":             "Product Name",
    "order_status":             "Order Status",
    "rrp_used":                 "Rrp Used",
    "srp_used":                 "Srp Used",
    "seller_srp_disc_pct":      "Seller SRP Discount %",
    "seller_vc_disc_pct":       "SELLER Voucher Discount % Mentioned in Exclusion Remark",
    "seller_end_disc_pct":      "SELLER END DISCOUNT %",
    "paid_price":               "Customer PAID Price",
    "actual_total_disc_pct":    "Customer Disc % From RRP",
    "platform_discount_amount": "Platform Discount Amount",
    "effective_price":          "Effective Price (Paid + Platform Disc)",
    "effective_disc_pct":       "Effective Disc % From RRP",
    "remark":                   "Remark",
    "rule_label":               "Rule Label",
    "rule_type":                "Rule Type",
    "flagged":                  "Flagged",
    "flag_reason":              "Flag Reason",
    "flag_severity":            "Flag Severity",
}


def build_report(result_df: pd.DataFrame) -> bytes:
    from discount_engine import flagged_orders

    buf = io.BytesIO()

    # ── Sheet 1: Exclusion Summary ─────────────────────────────────────────────
    excl, excl_sev = _build_excl_summary(result_df)

    # ── Sheet 2: Full Detail ───────────────────────────────────────────────────
    detail_cols = [c for c in DETAIL_COLS if c in result_df.columns]
    detail      = result_df[detail_cols].copy()
    detail.rename(columns=COL_LABELS, inplace=True)
    sev_series  = result_df["flag_severity"] if "flag_severity" in result_df.columns else pd.Series(dtype=str)

    # ── Sheet 3: Flagged Orders ────────────────────────────────────────────────
    flagged     = flagged_orders(result_df)
    flag_cols   = [c for c in DETAIL_COLS if c in flagged.columns]
    flagged_out = flagged[flag_cols].copy()
    flagged_out.rename(columns=COL_LABELS, inplace=True)
    flag_sev    = flagged["flag_severity"] if "flag_severity" in flagged.columns else pd.Series(dtype=str)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        excl.to_excel(writer, sheet_name="Exclusion Summary", index=False)
        _style_ws(writer.sheets["Exclusion Summary"], len(excl), excl_sev)

        detail.to_excel(writer, sheet_name="Full Detail", index=False)
        _style_ws(writer.sheets["Full Detail"], len(detail), sev_series)

        if not flagged_out.empty:
            flagged_out.to_excel(writer, sheet_name="Flagged Orders", index=False)
            _style_ws(writer.sheets["Flagged Orders"], len(flagged_out), flag_sev,
                      header_fill=RED_FILL)
        else:
            pd.DataFrame([{"Note": "No flagged orders ✅"}]).to_excel(
                writer, sheet_name="Flagged Orders", index=False)
            _style_ws(writer.sheets["Flagged Orders"], 1, pd.Series(dtype=str))

    return buf.getvalue()


def _build_excl_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Exclusion Summary columns (exact format file order):
      Region | Marketplace | Exclusion Remark | Rule | Orders |
      Sum of RRP | SUM OF PAID PRICE |
      Calculate the discount % from SUM OF PAID to SUM OF RRP |
      Violations | Status
    _severity column is excluded from output.
    """
    status_col = df.get("order_status", pd.Series("", index=df.index))
    active = df[~status_col.astype(str).str.lower().str.contains("cancel", na=False)].copy()

    remark_col = "remark"      if "remark"      in active.columns else None
    rule_col   = "rule_label"  if "rule_label"  in active.columns else None
    sev_col    = "flag_severity" if "flag_severity" in active.columns else None
    grp_keys   = [c for c in [remark_col, rule_col, sev_col, "region", "marketplace"] if c]

    rows = []
    if not grp_keys or active.empty:
        return pd.DataFrame()

    for keys, grp in active.groupby(grp_keys, dropna=False, observed=True):
        if not isinstance(keys, tuple):
            keys = (keys,)
        key_map  = dict(zip(grp_keys, keys))
        remark   = str(key_map.get(remark_col, ""))
        rule     = str(key_map.get(rule_col, ""))
        severity = str(key_map.get(sev_col, "grey"))
        region   = str(key_map.get("region", ""))
        mp       = str(key_map.get("marketplace", ""))

        sum_rrp  = grp["rrp_used"].sum()   if "rrp_used"   in grp.columns else 0
        sum_paid = grp["paid_price"].sum()  if "paid_price" in grp.columns else 0
        orders   = len(grp)
        flagged  = int(grp["flagged"].sum()) if "flagged" in grp.columns else 0

        # Disc % = (Sum RRP - Sum Paid) / Sum RRP × 100
        disc_pct = round((sum_rrp - sum_paid) / sum_rrp * 100, 1) if sum_rrp > 0 else 0.0

        rows.append({
            "Region":            region,
            "Marketplace":       mp,
            "Exclusion Remark":  remark if remark not in ("", "nan") else "(no remark)",
            "Rule":              rule,
            "Orders":            orders,
            "Sum of RRP":        round(sum_rrp, 2),
            "SUM OF PAID PRICE": round(sum_paid, 2),
            "Calculate the discount % from SUM OF PAID to SUM OF RRP": disc_pct,
            "Violations":        flagged,
            "Status":            "🚨 Violated" if flagged > 0 else "✅ OK",
            "_sev":              severity,   # keep for row colouring, hidden from output
        })

    if not rows:
        return pd.DataFrame()

    out = (pd.DataFrame(rows)
           .sort_values(["Violations", "Calculate the discount % from SUM OF PAID to SUM OF RRP"],
                        ascending=[False, False])
           .reset_index(drop=True))

    # Extract severity for row colouring BEFORE dropping
    sev_series_out = out["_sev"].copy().reset_index(drop=True)
    out.drop(columns=["_sev"], inplace=True)   # ← hidden from Excel output
    # Return as tuple (df, severity_series) — caller unpacks
    return out, sev_series_out


def _style_ws(ws, nrows: int, severities: pd.Series, header_fill=None):
    hf = header_fill or DARK_FILL
    for cell in ws[1]:
        cell.fill = hf
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    sev_list = list(severities.reset_index(drop=True)) if len(severities) > 0 else []
    for ri, row in enumerate(ws.iter_rows(min_row=2)):
        sev  = sev_list[ri] if ri < len(sev_list) else None
        fgc  = SEV_COLORS.get(str(sev), "")
        fill = PatternFill("solid", fgColor=fgc) if fgc else (ALT_FILL if ri % 2 == 0 else PatternFill())
        for cell in row:
            cell.fill      = fill
            cell.alignment = LEFT
